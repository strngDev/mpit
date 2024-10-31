import csv
from openpyxl import Workbook


def save_to_xlsx(data_list):
    wb = Workbook()
    ws = wb.active

    # Определяем заголовки столбцов
    headers = list(data_list[0].keys()) if len(data_list) > 0 else []
    # Заполняем первую строку заголовками
    for col in range(len(headers)):
        ws.cell(row=1, column=col + 1).value = headers[col]

    # Заполняем строки значениями
    row_num = 2
    for data in data_list:
        for col, value in enumerate(data.values()):
            ws.cell(row=row_num, column=col + 1).value = value
        row_num += 1

    # Сохраняем файл
    wb.save("output.xlsx")

def write_to_csv(data):
    with open('output.csv', mode='a', newline='', encoding='utf-8') as file:
        fieldnames = ['name', 'age', 'city', 'username', 'resume', 'spec']
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        for row in data:
            writer.writerow(row)


def read_from_csv() -> list:
    result = []
    with open('output.csv', mode='r', newline='', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            result.append(dict(row))
    return result